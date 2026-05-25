"""Exportación de Personal a Excel (.xlsx)."""

from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment as XlAlign, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.dto.Personal.personal_response_dto import PersonalResponseDTO

EXPORTABLE_COLUMNS: list[tuple[str, str]] = [
    ("# Empleado", "num_empleado"),
    ("Nombre completo", "_full_name"),
    ("Nombres", "nombres"),
    ("Apellido Paterno", "apellido_paterno"),
    ("Apellido Materno", "apellido_materno"),
    ("Correo corporativo", "mail"),
    ("Correo nómina", "correo_nomina"),
    ("Departamento", "nombre_departamento"),
    ("Área", "nombre_area"),
    ("Cargo", "nombre_tc"),
    ("Tipo de puesto", "nombre_tipo_puesto"),
    ("Jefe directo", "nombre_jefe"),
    ("Estado", "_status"),
    ("Rol App", "rol_app"),
    ("Contraseña", "password_plain"),
]


def _cell_value(item: PersonalResponseDTO, field: str) -> str:
    if field == "_full_name":
        parts = [
            getattr(item, "nombres", "") or "",
            getattr(item, "apellido_paterno", "") or "",
            getattr(item, "apellido_materno", "") or "",
        ]
        return " ".join(p for p in parts if p).strip()
    if field == "_status":
        return "Activo" if getattr(item, "activo", True) else "Inactivo"
    return str(getattr(item, field, "") or "")


def exportar_personal(
    items: Iterable[PersonalResponseDTO],
    columnas: list[tuple[str, str]],
    ruta: str | Path,
) -> Path:
    ruta = Path(ruta)
    wb = Workbook()
    ws = wb.active
    ws.title = "Personal"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1565C0")
    header_align = XlAlign(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, (label, _) in enumerate(columnas, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    ws.row_dimensions[1].height = 22

    for row_idx, item in enumerate(items, start=2):
        for col_idx, (_, field) in enumerate(columnas, start=1):
            ws.cell(row=row_idx, column=col_idx, value=_cell_value(item, field))

    for col_idx, (label, _) in enumerate(columnas, start=1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = max(len(label) + 4, 16)

    ws.freeze_panes = "A2"
    wb.save(ruta)
    return ruta
