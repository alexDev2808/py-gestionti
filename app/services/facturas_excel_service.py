"""Exportación de facturas a Excel.

Genera un archivo .xlsx por filial dentro de la carpeta destino configurada.
El archivo se reescribe completo cada vez para garantizar consistencia con la BD.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Iterable, Optional

from app.dto.Facturas.facturas_response_dto import FacturasResponseDTO

_HEADERS = [
    "ID", "Filial", "Proveedor", "Cliente",
    "Período", "Núm. factura", "Monto",
    "Estado", "Fecha descarga", "Fecha envío",
    "Destinatario", "Ruta PDF", "Ruta XML",
    "Notas", "Creado por", "Creado en", "Actualizado en",
]


def _fmt(d: Optional[datetime]) -> str:
    return d.strftime("%Y-%m-%d %H:%M") if d else ""


def _row(item: FacturasResponseDTO) -> list:
    return [
        item.id_factura,
        item.filial_nombre,
        item.proveedor_nombre,
        item.cliente_nombre,
        item.periodo,
        item.numero_factura,
        float(item.monto) if item.monto is not None else "",
        item.estado,
        _fmt(item.fecha_descarga),
        _fmt(item.fecha_envio),
        item.destinatario,
        item.ruta_pdf,
        item.ruta_xml,
        item.notas,
        item.creado_por,
        _fmt(item.creado_en),
        _fmt(item.actualizado_en),
    ]


class FacturasExcelService:
    """Exportador de facturas a archivos Excel por filial."""

    def __init__(self, base_folder: Optional[Path] = None) -> None:
        self.base_folder = Path(base_folder) if base_folder else Path("C:/GestionTI/Facturas")

    def _ensure_folder(self) -> Path:
        self.base_folder.mkdir(parents=True, exist_ok=True)
        return self.base_folder

    def export_filial(
        self,
        filial_nombre: str,
        facturas: Iterable[FacturasResponseDTO],
    ) -> Path:
        """
        Genera/sobrescribe el Excel de una filial completa.

        Argumentos:
            filial_nombre (str): Nombre de la filial (México, Tlaxcala).
            facturas (Iterable[FacturasResponseDTO]): Facturas a exportar.

        Retorna:
            Path: Ruta del archivo generado.

        Lanza:
            ImportError: Si openpyxl no está disponible.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError as exc:  # pragma: no cover
            raise ImportError(
                "openpyxl no está instalado. Agrégalo con: pip install openpyxl"
            ) from exc

        carpeta = self._ensure_folder()
        slug = "".join(c for c in filial_nombre if c.isalnum() or c in (" ", "-", "_")).strip().replace(" ", "_")
        nombre_archivo = f"Facturas_{slug or 'filial'}.xlsx"
        destino = carpeta / nombre_archivo

        wb = Workbook()
        ws = wb.active
        ws.title = filial_nombre[:31] or "Facturas"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_align = Alignment(horizontal="center", vertical="center")

        for col_idx, header in enumerate(_HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        for row_idx, item in enumerate(facturas, start=2):
            for col_idx, value in enumerate(_row(item), start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-ajustar anchos
        for col_idx, header in enumerate(_HEADERS, start=1):
            max_len = max(
                [len(str(header))] +
                [len(str(ws.cell(row=r, column=col_idx).value or ""))
                 for r in range(2, ws.max_row + 1)]
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

        ws.freeze_panes = "A2"
        wb.save(destino)
        return destino

    def export_todas(
        self,
        por_filial: dict[str, list[FacturasResponseDTO]],
    ) -> list[Path]:
        """Genera un archivo por cada filial presente en el diccionario."""
        return [self.export_filial(nombre, items) for nombre, items in por_filial.items()]
