"""Servicio para importar correos de nómina desde un Excel.

Espera un archivo .xlsx con encabezados literales en la primera fila:
    num_empleado, email
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Optional


@dataclass
class FilaImport:
    fila: int                    # Fila del Excel (1-indexed con encabezado en fila 1)
    num_empleado: str
    correo: Optional[str]
    estado: str                  # 'ok' | 'sin_correo' | 'invalido' | 'no_existe'
    motivo: str = ""


@dataclass
class ResumenImport:
    filas: list[FilaImport]
    total: int
    listos: int
    sin_correo: int
    invalidos: int
    no_existen: int


def _es_correo_valido(s: str) -> bool:
    if not s or "@" not in s:
        return False
    user, _, dom = s.rpartition("@")
    return bool(user) and "." in dom


def leer_excel(path: str | Path) -> list[FilaImport]:
    """Lee el Excel y devuelve filas crudas (sin chequear contra BD).

    Lanza ImportError si openpyxl no está, FileNotFoundError si el archivo no existe,
    ValueError si los encabezados no coinciden.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError(
            "openpyxl no está instalado. Agrégalo con: pip install openpyxl"
        ) from exc

    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"Archivo no encontrado: {p}")

    wb = load_workbook(p, read_only=True, data_only=True)
    ws = wb.active

    # Encabezados en fila 1
    encabezados = [
        (str(c.value).strip().lower() if c.value is not None else "")
        for c in next(ws.iter_rows(min_row=1, max_row=1))
    ]
    try:
        idx_num = encabezados.index("num_empleado")
        idx_email = encabezados.index("email")
    except ValueError:
        wb.close()
        raise ValueError(
            "El Excel debe tener encabezados literales 'num_empleado' y 'email' en la fila 1."
        )

    filas: list[FilaImport] = []
    for fila_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # row es una tupla; protegerse de filas más cortas
        num_raw = row[idx_num] if idx_num < len(row) else None
        email_raw = row[idx_email] if idx_email < len(row) else None

        num = str(num_raw).strip() if num_raw is not None else ""
        if not num:
            # Salta filas vacías
            continue

        correo = str(email_raw).strip() if email_raw is not None else ""

        if not correo:
            filas.append(FilaImport(
                fila=fila_idx, num_empleado=num, correo=None,
                estado="sin_correo", motivo="Sin correo en el Excel",
            ))
        elif not _es_correo_valido(correo):
            filas.append(FilaImport(
                fila=fila_idx, num_empleado=num, correo=correo,
                estado="invalido", motivo="Formato de correo inválido",
            ))
        else:
            filas.append(FilaImport(
                fila=fila_idx, num_empleado=num, correo=correo,
                estado="ok",
            ))

    wb.close()
    return filas


def construir_resumen(filas: list[FilaImport], existentes: set[str]) -> ResumenImport:
    """Marca las filas que no existen en BD y arma el resumen agregado."""
    for f in filas:
        if f.estado == "ok" and f.num_empleado not in existentes:
            f.estado = "no_existe"
            f.motivo = "Empleado no encontrado en Personal"

    listos = sum(1 for f in filas if f.estado == "ok")
    sin_correo = sum(1 for f in filas if f.estado == "sin_correo")
    invalidos = sum(1 for f in filas if f.estado == "invalido")
    no_existen = sum(1 for f in filas if f.estado == "no_existe")

    return ResumenImport(
        filas=filas,
        total=len(filas),
        listos=listos,
        sin_correo=sin_correo,
        invalidos=invalidos,
        no_existen=no_existen,
    )


def filas_para_aplicar(resumen: ResumenImport) -> list[tuple[str, Optional[str]]]:
    """Solo las filas 'ok' se persisten — las inválidas/sin_correo/no_existe se descartan."""
    return [(f.num_empleado, f.correo) for f in resumen.filas if f.estado == "ok"]
